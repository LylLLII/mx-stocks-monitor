"""
妙想选股 - 盘中累计观察池监控脚本（自包含 + 自带授权，可跨机器拷贝）
================================================================
依赖: 仅 httpx（可选 qrcode 用于生成授权二维码）
不依赖任何外部技能目录；首次运行会自动引导扫码授权，之后免授权。

家用电脑部署步骤（极简）:
  1) 把本文件拷到家里电脑任意目录
  2) 装依赖:  pip install httpx qrcode
  3) 运行一次:  python screener_monitor.py --force
       -> 首次会打印/生成授权二维码，手机扫码授权
  4) 再运行一次:  python screener_monitor.py --force   (此时已授权，开始工作)
  5) 建一个同样的定时任务（工作日每2分钟，脚本自动守卫交易时段）

用法:
  python screener_monitor.py            # 仅交易时段执行，命中即累计合并到观察池
  python screener_monitor.py --force    # 忽略交易时段守卫（用于手动建池/测试/首次授权）
"""
import argparse
import asyncio
import csv
import json
import os
import stat
import sys
import uuid
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

import httpx  # pip install httpx

# WORKSPACE 跟随脚本自身位置，clone 到任何目录/机器都能用
WORKSPACE = Path(__file__).resolve().parent
OUTPUT_DIR = WORKSPACE / "mx_stocks_screener"
TMP_DIR = OUTPUT_DIR / "_tmp"
MASTER_CSV = OUTPUT_DIR / "观察池_累计.csv"

MCP_URL = "https://ai-saas.eastmoney.com/proxy/b/mcp/tool/selectSecurity"
AUTH_BASE = os.environ.get("EM_AUTH_BASE", "https://ai-saas.eastmoney.com").rstrip("/")
CLIENT_ID = "mx-stocks-screener"
API_KEY_PAGE_URL = "https://ai.eastmoney.com/mxClaw"
DEFAULT_EXPIRES_IN = 30 * 24 * 60 * 60

# 自定义异常：用于区分「云端缺密钥」与「key 过期」两类失败，便于上层精准告警
class MissingSecret(RuntimeError):
    """云端运行却拿不到 EM_API_KEY（仓库未配置 Secret）。"""


class KeyExpired(RuntimeError):
    """EM_API_KEY 已失效（HTTP 401 或业务码 401）。"""


# 本地 --no-push 开关：置 True 时 git_sync_after 只 pull、不 commit/push，
# 避免本地与云端（唯一写入方）抢同一份 CSV。
_NO_PUSH = False


# ---------------- 自带授权（无需外部技能） ----------------
def _mx_dir() -> Path:
    return Path.home() / ".mx-skills"


def _key_path() -> Path:
    return _mx_dir() / "em_api_key"


def _pending_path() -> Path:
    return _mx_dir() / "pending_auth.json"


def _has_valid_key() -> bool:
    if (os.environ.get("EM_API_KEY") or "").strip():
        return True
    p = _key_path()
    return p.exists() and bool(p.read_text(encoding="utf-8").strip())


def _load_api_key() -> str:
    v = (os.environ.get("EM_API_KEY") or "").strip()
    if v:
        return v
    p = _key_path()
    if p.exists():
        return p.read_text(encoding="utf-8").strip()
    return ""


def _write_api_key(api_key: str) -> None:
    p = _key_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(api_key.strip() + "\n", encoding="utf-8")
    try:
        os.chmod(p, stat.S_IRUSR | stat.S_IWUSR)
    except (OSError, NotImplementedError):
        pass


def _read_pending():
    p = _pending_path()
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        _clear_pending()
        return None
    if not isinstance(data, dict) or not data.get("token"):
        _clear_pending()
        return None
    if int(data.get("expiresAt", 0)) <= int(__import__("time").time()):
        _clear_pending()
        return None
    return data


def _write_pending(token, auth_url, api_key_url, expires_in):
    p = _pending_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps({
        "token": token, "authUrl": auth_url, "apiKeyUrl": api_key_url,
        "expiresAt": int(__import__("time").time()) + int(expires_in) - 5,
    }, ensure_ascii=False), encoding="utf-8")
    try:
        os.chmod(p, stat.S_IRUSR | stat.S_IWUSR)
    except (OSError, NotImplementedError):
        pass


def _clear_pending() -> None:
    p = _pending_path()
    if p.exists():
        try:
            p.unlink()
        except OSError:
            pass


def _api_create():
    r = httpx.post(f"{AUTH_BASE}/api/auth/token/create", json={"clientId": CLIENT_ID}, timeout=30.0)
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, "0", 200, "200", None):
        raise RuntimeError(d.get("message") or d.get("msg") or "create 业务错误")
    data = d.get("data")
    if not data or not data.get("token") or not data.get("authUrl"):
        raise RuntimeError(f"create 未返回 token/authUrl: {data}")
    return data["token"], data["authUrl"], data.get("apiKeyUrl") or API_KEY_PAGE_URL, int(data.get("expiresIn", DEFAULT_EXPIRES_IN))


def _api_result(token):
    r = httpx.post(f"{AUTH_BASE}/api/auth/token/result", json={"token": token}, timeout=30.0)
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, "0", 200, "200", None):
        raise RuntimeError(d.get("message") or d.get("msg") or "result 业务错误")
    data = d.get("data") or {}
    return data.get("state") or "invalid", data.get("apiKey")


def _print_auth(auth_url, api_key_url):
    print("=" * 56)
    print("首次使用需要授权（扫码一次即可，之后免授权）:")
    print(f"  扫码授权链接: {auth_url}")
    print(f"  或手动复制 apikey: {api_key_url}")
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(auth_url)
        qr.make(fit=True)
        out = Path("auth_qr.png")
        qr.make_image(fill_color="black", back_color="white").save(out)
        print(f"  二维码已生成: {out.resolve()}  (用手机扫码)")
    except Exception:
        print("  (未安装 qrcode，可直接打开上面的链接扫码)")
    print("扫码完成后，重新运行一次本脚本即可开始监控。")
    print("=" * 56)


def ensure_api_key() -> str:
    """返回可用 key；若无则引导授权并退出，下次运行自动落盘。"""
    if _has_valid_key():
        return _load_api_key()
    # 云端（无头）环境拿不到本地 key 文件、也没有 EM_API_KEY 环境变量：
    # 必须显式失败（红色 ✗），绝不能 sys.exit(0) 制造「成功」假象。
    if os.environ.get("GITHUB_ACTIONS") == "true":
        raise MissingSecret(
            "云端缺少 EM_API_KEY：请在仓库 Settings → Secrets and variables → Actions "
            "新建 Secret（Name=EM_API_KEY，Value=本地 ~/.mx-skills/em_api_key 的内容，去掉换行）。"
        )
    pending = _read_pending()
    if pending is not None:
        try:
            state, api_key = _api_result(pending["token"])
        except Exception:
            _clear_pending()
            pending = None
        if pending is not None:
            if state == "done" and api_key:
                _write_api_key(api_key)
                _clear_pending()
                return api_key
            if state == "pending":
                _print_auth(pending.get("authUrl") or pending.get("qrUrl"), pending.get("apiKeyUrl") or API_KEY_PAGE_URL)
                sys.exit(0)
            _clear_pending()
    token, auth_url, api_key_url, exp = _api_create()
    _write_pending(token, auth_url, api_key_url, exp)
    _print_auth(auth_url, api_key_url)
    sys.exit(0)


# ---------------- 续期 / 过期告警 ----------------
def _notify_key_expired():
    """key 失效时通知：云端开一个去重 Issue；本地打印醒目提示。"""
    if os.environ.get("GITHUB_ACTIONS") == "true":
        _create_expiry_issue()
    else:
        print("=" * 56)
        print("⚠️  EM_API_KEY 已失效！选股已暂停。")
        print("请在本机运行一次续期（约 30 秒扫码）：")
        print("    python screener_monitor.py --renew")
        print("续期后脚本会自动把新 key 推送到仓库 Secret，云端下个周期自动恢复。")
        print("=" * 56)


def _dedupe_issue_title() -> str:
    return "🔑 EM_API_KEY 已过期 — 请本地运行 --renew 续期"


def _create_expiry_issue():
    """在云端仓库开一个去重的续期提醒 Issue（优先用 gh，回退 REST）。"""
    repo = os.environ.get("GITHUB_REPOSITORY")
    if not repo:
        print("[告警] 无法获取 GITHUB_REPOSITORY，跳过创建续期 Issue")
        return
    title = _dedupe_issue_title()
    body = (
        "东方财富 API key 已失效（HTTP 401），选股已暂停。\n\n"
        "请在本机执行：\n"
        "    python screener_monitor.py --renew\n"
        "扫码授权后，脚本会自动把新 key 推送到仓库 Secret（EM_API_KEY），"
        "云端下个触发周期会自动用上新 key，无需其他操作。"
    )
    try:
        import subprocess
        # Windows 下抑制子进程控制台窗口（gh.exe 弹窗同样烦人）
        creationflags = 0
        if os.name == "nt":
            creationflags = subprocess.CREATE_NO_WINDOW
        # 去重：先看是否已有未关闭的同名 Issue
        r = subprocess.run(
            ["gh", "issue", "list", "--repo", repo, "--state", "open",
             "--label", "key-expiry", "--json", "title"],
            capture_output=True, text=True, timeout=30, creationflags=creationflags)
        if r.returncode == 0:
            import json as _j
            for it in _j.loads(r.stdout or "[]"):
                if title in (it.get("title") or ""):
                    print("[告警] 已存在未关闭的续期 Issue，跳过重复创建")
                    return
        r2 = subprocess.run(
            ["gh", "issue", "create", "--repo", repo, "--title", title,
             "--body", body, "--label", "key-expiry"],
            capture_output=True, text=True, timeout=30, creationflags=creationflags)
        if r2.returncode == 0:
            print(f"[告警] 已创建续期 Issue: {r2.stdout.strip()}")
            return
        print(f"[告警] gh issue create 失败: {r2.stderr.strip()[:200]}")
    except Exception as e:
        print(f"[告警] 创建 Issue 异常: {type(e).__name__}: {e}")
    # 回退：REST API（需要 GITHUB_TOKEN，Actions 已注入）
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        return
    try:
        headers = {"Authorization": f"Bearer {token}",
                   "Accept": "application/vnd.github+json"}
        rr = httpx.get(
            f"https://api.github.com/repos/{repo}/issues",
            params={"state": "open", "labels": "key-expiry", "per_page": 20},
            headers=headers, timeout=20)
        if rr.status_code == 200:
            for it in rr.json():
                if title in (it.get("title") or ""):
                    print("[告警] 已存在未关闭的续期 Issue，跳过")
                    return
        cr = httpx.post(
            f"https://api.github.com/repos/{repo}/issues", headers=headers,
            json={"title": title, "body": body, "labels": ["key-expiry"]}, timeout=20)
        print(f"[告警] REST 创建 Issue 状态: {cr.status_code}")
    except Exception as e:
        print(f"[告警] REST 创建 Issue 异常: {type(e).__name__}: {e}")


def _push_secret(api_key: str):
    """把新 key 推送到仓库 Secret EM_API_KEY（需本地有 gh 且已登录 / 能推断仓库）。"""
    repo = os.environ.get("GITHUB_REPOSITORY")
    if not repo:
        try:
            import subprocess
            r = subprocess.run(["git", "config", "--get", "remote.origin.url"],
                               cwd=str(WORKSPACE), capture_output=True,
                               text=True, timeout=20)
            url = (r.stdout or "").strip()
            m = re.search(r"github\.com[:/]([^/]+/[^/.]+)", url)
            if m:
                repo = m.group(1)
        except Exception:
            pass
    if not repo:
        print("[续期] 未能确定仓库，请手动在仓库 Settings→Secrets 添加 EM_API_KEY。")
        return
    try:
        import subprocess
        r = subprocess.run(
            ["gh", "secret", "set", "EM_API_KEY", "--repo", repo, "--body", api_key.strip()],
            capture_output=True, text=True, timeout=60)
        if r.returncode == 0:
            print(f"[续期] 已自动推送新 key 到仓库 {repo} 的 Secret EM_API_KEY，"
                  f"云端下个周期自动恢复。")
        else:
            print(f"[续期] 自动推送失败({r.stderr.strip()[:200]})，"
                  f"请手动在仓库 Settings→Secrets 添加 EM_API_KEY。")
    except FileNotFoundError:
        print("[续期] 本机未安装 gh，请手动在仓库 Settings→Secrets 添加 EM_API_KEY"
              f"（值为新 key，去换行）。")
    except Exception as e:
        print(f"[续期] 推送异常: {type(e).__name__}: {e}")


def cmd_renew():
    """本地续期命令：测试当前 key → 失效则打印二维码轮询 → 写入并推送新 key。"""
    print("=== 续期 EM_API_KEY ===")
    cur = _load_api_key()
    if cur:
        try:
            asyncio.run(mcp_call(QUERY, "A股", cur))
            print("[续期] 当前 key 仍有效，无需续期。")
            return
        except KeyExpired:
            print("[续期] 当前 key 已失效，开始续期流程 ...")
        except Exception as e:
            print(f"[续期] 当前 key 测试异常({type(e).__name__})，为稳妥起见仍执行续期 ...")
    # 续期流程：复用 pending（若有未完成授权）或新建 token
    pending = _read_pending()
    if pending is None:
        token, auth_url, api_key_url, exp = _api_create()
        _write_pending(token, auth_url, api_key_url, exp)
    else:
        token = pending["token"]
        auth_url = pending.get("authUrl")
        api_key_url = pending.get("apiKeyUrl") or API_KEY_PAGE_URL
    _print_auth(auth_url, api_key_url)
    # 轮询授权结果（最多等 10 分钟，给人扫码时间）
    import time as _t
    deadline = _t.monotonic() + 600
    new_key = None
    while _t.monotonic() < deadline:
        try:
            state, api_key = _api_result(token)
        except Exception as e:
            print(f"[续期] 查询失败: {e}")
            _t.sleep(5)
            continue
        if state == "done" and api_key:
            new_key = api_key
            break
        print(f"[续期] 等待扫码授权 ... (state={state})")
        _t.sleep(5)
    if not new_key:
        print("[续期] 超时未授权，请扫码后再次运行 --renew。")
        sys.exit(1)
    _write_api_key(new_key)
    _clear_pending()
    print("[续期] 新 key 已写入本地 ~/.mx-skills/em_api_key")
    _push_secret(new_key)


# ---------------- 选股与累计逻辑 ----------------
QUERY = (
    "A股，剔除北交所、科创板、创业板，剔除ST股和退市股，剔除有新规风险的股票，"
    "剔除近三个月有减持计划的股票，剔除未来三个月有解禁的股票，剔除近三个月收到监管函或监管工作函的股票，"
    "剔除最新报告期净利润为负的股票，剔除资产负债率超过70%的股票，"
    "涨跌幅在3%到5%之间，量比大于1.5，换手率在5%到10%之间，总市值在50亿到200亿之间，近20日有涨停"
)

# 注：「最新价(元)」「涨跌幅(%)」仅用于首次入选时生成冻结快照（入选价/入选时涨跌幅），
# 本身不写入 CSV 输出列——避免「实时价」与「冻结基准」并存造成混淆（用户只需入选快照）。
_LIVE_ONLY = {"最新价(元)", "涨跌幅(%)"}

MAPPING = [
    ("代码", "代码"),
    ("名称", "名称"),
    ("上市板块", "上市板块"),
    ("最新价(元)", "最新价(元)"),   # 仅内部：首次入选时用于生成 入选价(元)
    ("涨跌幅(%)", "涨跌幅(%)"),       # 仅内部：首次入选时用于生成 入选时涨跌幅(%)
    ("量比", "量比"),
    ("换手率(%)", "换手率(%)"),
    ("总市值", "总市值(元)"),
    ("涨停次数(近20日)", "涨停次数(次)"),
    ("ST股票", "ST股票"),
    ("退市股", "退市股"),
    ("戴帽预期(新规)", "戴帽预期(新规)"),
]
T1_COLS = [
    "次日_跟踪状态", "次日_跟踪日期",
    "次日_开盘涨跌幅", "次日_午间涨跌幅", "次日_收盘涨跌幅", "次日_当前涨跌幅",
    "次日_最高涨跌幅", "次日_最低涨跌幅", "次日_形态",
    # 策略口径收益：以「入选价」为买入基准（信号出现时价格），而非昨收
    "策略收益_次日收盘(%)", "策略收益_次日最高(%)", "策略收益_次日最低(%)",
]
# 输出列 = MAPPING 中除「仅内部」字段外的所有列 + 冻结快照 + T1 跟踪
MASTER_COLS = [m[0] for m in MAPPING if m[0] not in _LIVE_ONLY] + [
    "首次入选日期", "首次入选时间", "入选价(元)", "入选时涨跌幅(%)",
    "入选次数", "入选扫描时间点"
] + T1_COLS


def now_shanghai() -> datetime:
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Asia/Shanghai"))
    except Exception:
        return (datetime.now(timezone.utc) + timedelta(hours=8)).replace(tzinfo=None)


def in_trading_hours(dt: datetime) -> bool:
    # 先判交易日：跳过周末与法定节假日（chinese_calendar 不可用时退化为仅判周末）
    try:
        import chinese_calendar as cn
        if not cn.is_workday(dt.date()):
            return False
    except Exception:
        if dt.weekday() >= 5:
            return False
    t = dt.time()
    morning = datetime.strptime("09:30", "%H:%M").time() <= t <= datetime.strptime("11:30", "%H:%M").time()
    afternoon = datetime.strptime("13:00", "%H:%M").time() <= t <= datetime.strptime("15:00", "%H:%M").time()
    return morning or afternoon


async def mcp_call(query: str, select_type: str, api_key: str) -> dict:
    meta = {
        "query": query,
        "selectType": select_type,
        "toolContext": {
            "callId": f"call_{uuid.uuid4().hex[:8]}",
            "userInfo": {"userId": f"user_{uuid.uuid4().hex[:8]}"},
        },
    }
    async with httpx.AsyncClient(timeout=30.0) as client:
        r = await client.post(
            MCP_URL, json=meta,
            headers={
                "Content-Type": "application/json",
                "em_api_key": api_key,
                "x-open-id-vendor": "tencent",
                "x-open-id-app": "workbuddy",
            },
        )
        if r.status_code == 401:
            raise KeyExpired("EM_API_KEY 失效 (HTTP 401)，需重新授权")
        try:
            payload = r.json()
        except Exception as e:
            raise RuntimeError(f"响应 JSON 解析失败: {e}") from e
        if isinstance(payload, dict):
            code = payload.get("code")
            status = payload.get("status")
            if code in (401, "401") or status in (401, "401"):
                raise KeyExpired("EM_API_KEY 失效 (业务码 401)，需重新授权")
            data = payload.get("data") or {}
            if isinstance(data, dict) and data:
                return data
        raise RuntimeError("MCP 返回为空或非预期结构")


def parse_rows(data: dict):
    result_node = (data.get("allResults") or {}).get("result") or {}
    data_list = result_node.get("dataList") or []
    columns = result_node.get("columns") or []
    if not data_list:
        return []
    col_map, order = {}, []
    for c in columns:
        if not isinstance(c, dict):
            continue
        en = c.get("field") or c.get("name") or c.get("key")
        cn = c.get("displayName") or c.get("title") or c.get("label") or en
        if c.get("dateMsg"):
            cn = f"{cn} {c['dateMsg']}"
        if en is not None and cn is not None:
            col_map[str(en)] = str(cn)
            order.append(str(en))
    rows = []
    for row in data_list:
        if not isinstance(row, dict):
            continue
        cn_row = {}
        for en in order:
            if en in row:
                v = row[en]
                cn_row[col_map[en]] = "" if v is None else (json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v))
        rows.append(cn_row)
    return rows


def find_col(headers, key):
    for h in headers:
        if h == key or h.startswith(key):
            return h
    return None


def run_screener() -> list:
    api_key = ensure_api_key()
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    data = asyncio.run(mcp_call(QUERY, "A股", api_key))
    rows = parse_rows(data)
    if not rows:
        return []
    headers = list(rows[0].keys())
    src_cols = {m[1]: find_col(headers, m[1]) for m in MAPPING}
    out = []
    for r in rows:
        code = (r.get(src_cols["代码"]) or "").strip()
        if not code:
            continue
        m = {}
        for mcol, key in MAPPING:
            col = src_cols[key]
            m[mcol] = (r.get(col) or "").strip() if col else ""
        out.append(m)
    return out


# ---------------- 基本面二次校验（排雷，不依赖妙想"自觉"） ----------------
# 妙想 QUERY 里写了"剔除亏损/高负债"，但 AI 接口未必严格执行、且结果不可审计。
# 这里在代码层对候选做硬校验：TTM 归母净利润 < 0 或 资产负债率 > 70% 直接拦截。
# 数据源：东财 F10 主财务指标（datacenter.eastmoney.com，无鉴权、字段稳定）。
# 口径：净利润用 TTM（拉 5 期滚动计算；算不出时降级最新一期累计）；
#       负债率用最新报告期（ZCFZL 字段，直接给 %）。
# 失败策略：接口异常 → 放行（best-effort，避免误杀，保持监控连续性），打日志。

_FUND_MAX_DEBT_RATIO = 70.0      # 资产负债率阈值 %
_FUND_CACHE = {}                 # code -> (日期, ttm净利, 负债率, 报告期)


def _secid(code: str) -> str:
    """6 开头→SH，其余→SZ（本项目 QUERY 已剔除科创板/创业板/北交所）。"""
    code = (code or "").strip()
    return f"{code}.SH" if code.startswith("6") else f"{code}.SZ"


def _fetch_fundamental(code: str):
    """拉取最新 5 期财务，返回 (ttm_netprofit, debt_ratio, report_date)；失败返回 None。"""
    today = datetime.now().strftime("%Y-%m-%d")
    cached = _FUND_CACHE.get(code)
    if cached and cached[0] == today:
        return cached[1], cached[2], cached[3]
    secid = _secid(code)
    url = ("https://datacenter.eastmoney.com/securities/api/data/v1/get"
           "?reportName=RPT_F10_FINANCE_MAINFINADATA"
           "&columns=SECUCODE,REPORT_DATE,PARENTNETPROFIT,ZCFZL"
           f"&filter=(SECUCODE%3D%22{secid}%22)"
           "&pageNumber=1&pageSize=5&sortTypes=-1&sortColumns=REPORT_DATE")
    try:
        r = httpx.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        rows = ((r.json().get("result") or {}).get("data")) or []
    except Exception as e:
        print(f"[基本面] {code} 数据获取失败({type(e).__name__})，放行")
        return None
    if not rows:
        print(f"[基本面] {code} 无财务数据，放行")
        return None

    # 解析各期 (year, month, 净利, 负债率)
    periods = []
    for row in rows:
        rd = (row.get("REPORT_DATE") or "")[:10]
        try:
            year, month = int(rd[:4]), int(rd[5:7])
        except (ValueError, TypeError):
            continue
        np_ = row.get("PARENTNETPROFIT")
        debt = row.get("ZCFZL")
        if np_ is None:
            continue
        periods.append((year, month, float(np_), float(debt) if debt is not None else None))

    if not periods:
        print(f"[基本面] {code} 财务数据不可解析，放行")
        return None
    periods.sort(key=lambda x: (x[0], x[1]))  # 时间升序

    p0 = periods[-1]                     # 最新一期
    y0, m0, np0, debt0 = p0
    # TTM：最新期累计 + 最近年报 - 去年同期累计；若最新期即年报或期数不足 → 降级最新一期
    ttm = np0
    if m0 != 12:
        annual = next((p for p in reversed(periods)
                       if p[1] == 12 and (p[0], p[1]) < (y0, m0)), None)
        same_last = next((p for p in periods if p[1] == m0 and p[0] < y0), None)
        if annual and same_last:
            ttm = np0 + annual[2] - same_last[2]
    _FUND_CACHE[code] = (today, ttm, debt0, f"{y0}-{m0:02d}")
    return ttm, debt0, f"{y0}-{m0:02d}"


def _fundamental_gate(rows):
    """对候选 rows（list[dict]，含"代码"列）做基本面硬校验，返回过滤后的列表。"""
    if not rows:
        return rows
    kept, blocked = [], []
    for m in rows:
        code = (m.get("代码") or "").strip()
        if not code:
            kept.append(m)
            continue
        fd = _fetch_fundamental(code)
        if fd is None:
            kept.append(m)  # 数据拿不到 → 放行
            continue
        ttm, debt, rep = fd
        reasons = []
        if ttm is not None and ttm < 0:
            reasons.append(f"TTM净利润 {ttm/1e4:.0f} 万")
        if debt is not None and debt > _FUND_MAX_DEBT_RATIO:
            reasons.append(f"负债率 {debt:.1f}%")
        if reasons:
            blocked.append((code, m.get("名称"), "; ".join(reasons), rep))
            print(f"[基本面][拦截] {code} {m.get('名称')}（{rep}）：{'；'.join(reasons)}")
        else:
            kept.append(m)
    if blocked:
        print(f"[基本面] 拦截 {len(blocked)} 只: {[b[0] for b in blocked]} | 放行 {len(kept)} 只")
    return kept


def load_master() -> dict:
    pool = {}
    if MASTER_CSV.exists():
        # 注意：CSV 可能带 UTF-8 BOM（Excel 另存为会加 ef bb bf），
        # 必须用 utf-8-sig 打开，否则首列名变成 '\ufeff代码' 导致所有行被跳过、观察池被清空。
        with MASTER_CSV.open(encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                code = (r.get("代码") or "").strip()
                if not code:
                    continue
                # 防御：跳过历史遗留的日期分隔标记行（#===== 2026-08-05 =====）。
                # 当前 CSV 已不再写此类行（会破坏列一致性、导致 GitHub 表格视图消失），
                # 此处仅兼容旧文件，新数据不会触发。
                if code.startswith("#"):
                    continue
                # 防 git 冲突标记固化：pull/merge 冲突残留的 <<<<<<< HEAD / ======= / >>>>>>> 行
                # 会被 DictReader 解析成"代码"字段，若不拦截会被 _write_pool 原样写回并 commit 入库。
                if code.startswith(("<<<<<<<", "=======", ">>>>>>>")):
                    print(f"[严重] 检测到 git 冲突标记行，已跳过（{code[:30]}...）。"
                          f"工作区可能残留未解决的合并冲突，请检查。")
                    continue
                # 复合键 = 代码 + 首次入选日期：同一只票跨日期重新入选时各自成行（按批次独立跟踪 T+1）。
                fd = (r.get("首次入选日期") or "").strip()
                pool[f"{code}|{fd}" if fd else code] = r
    return pool


def _git(args):
    """best-effort git 调用，失败仅提示，不中断监控。"""
    try:
        import subprocess
        # Windows 下抑制子进程控制台窗口（避免每 3 分钟弹一次 git.exe 黑窗）
        creationflags = 0
        if os.name == "nt":
            creationflags = subprocess.CREATE_NO_WINDOW
        r = subprocess.run(["git"] + args, cwd=str(WORKSPACE),
                           capture_output=True, text=True, encoding="utf-8",
                           errors="replace", timeout=60,
                           creationflags=creationflags)
        if r.returncode != 0:
            msg = (r.stderr or r.stdout).strip().replace("\n", " ")
            print(f"[git] 跳过: {' '.join(args)} -> {msg[:160]}")
        return r.returncode == 0
    except Exception as e:
        print(f"[git] 跳过: {type(e).__name__}: {e}")
        return False


def git_sync_before() -> bool:
    """拉取远端最新观察池。失败返回 False——调用方必须跳过本轮（不扫描、不写盘），
    否则会基于旧数据扫描，且 pull 冲突残留的标记可能被 load_master 读入并固化。
    用 pull --rebase（而非 merge）与 git_sync_after 保持同一套冲突处理语义。"""
    return _git(["pull", "--rebase", "--no-edit"])


def git_sync_after():
    """推送本次增量到远端。push 失败 → pull --rebase → 重推一次；仍失败则 abort rebase 并留本地。"""
    if _NO_PUSH:
        # 本地镜像模式：绝不 commit/push，避免与云端（唯一写入方）抢同一份 CSV
        print("[git] --no-push：跳过 commit/push，仅保留本地镜像（数据以云端为准）")
        return True
    _git(["add", str(MASTER_CSV)])
    if MASTER_XLSX.exists():
        _git(["add", str(MASTER_XLSX)])
    if not _git(["commit", "-m", f"local: 更新观察池 {now_shanghai():%Y-%m-%d %H:%M}"]):
        # nothing to commit 也算成功
        return True
    if _git(["push"]):
        return True
    # push 被拒：远端有新的 commit（如云端同期推送）。先 pull --rebase 解冲突，再重推
    print("[git] push 被拒，尝试 pull --rebase 合并远端增量后重推 ...")
    if _git(["pull", "--rebase", "--no-edit"]) and _git(["push"]):
        print("[git] push-rebase-push 成功：本地增量已叠加到远端最新状态之上")
        return True
    # rebase 仍冲突 / push 仍失败：abort rebase 保住本地工作区，留待下次运行重试
    print("[git] rebase+push 持续失败，abort rebase 保住本地增量（本次未推送，不会丢数据）")
    _git(["rebase", "--abort"])
    return False


def _merge_rows(pool, rows, now):
    """把本次命中合并进 pool（in-memory），返回 (added, updated) 代码列表。
    语义：同一只票「同一日」多次命中 → 合并到当日行（刷新行情 + 计数 + 追加时间点）；
    跨日期再次命中（如 8.5 入选后 8.7 又入选）→ 按「代码+入选日期」新开一行，各自独立跟踪 T+1。
    历史行（之前日期的批次）保持冻结，仅由 track_followups 更新其 T+1 表现字段。"""
    ts = now.strftime("%Y-%m-%d %H:%M")
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M")
    added, updated = [], []
    for m in rows:
        code = m["代码"]
        today_key = f"{code}|{date}"
        if today_key not in pool:
            # 今天首次命中（含跨日期重新入选）：新开一行，冻结今日入选快照
            snap_price, snap_pct = _snapshot_for_new(m)
            new = dict(m)
            new.update({
                "首次入选日期": date, "首次入选时间": time,
                # 冻结本批次入选快照：以腾讯实时价（信号出现时）为基准，本行后续命中不再覆盖；
                # 腾讯不可达时回退妙想最新价（见 _snapshot_for_new）。
                # 注意：仅「入选价/入选时涨跌幅/首次入选日期+时间」冻结，
                # 其余行情字段（量比/换手率/总市值等）当日重复命中会刷新。
                "入选价(元)": snap_price, "入选时涨跌幅(%)": snap_pct,
                "入选次数": "1", "入选扫描时间点": ts,
            })
            pool[today_key] = new
            added.append(code)
        else:
            # 今日已命中过：仅刷新当日行的行情快照字段（MAPPING 列出的量比/换手率等），
            # 不触碰本行冻结的入选时间、入选价；历史日期行（旧批次）不更新。
            ex = pool[today_key]
            ex.update(m)
            ex["入选次数"] = str(int(ex.get("入选次数", "0") or 0) + 1)
            pts = [p for p in (ex.get("入选扫描时间点") or "").split(";") if p]
            pts.append(ts)
            ex["入选扫描时间点"] = ";".join(pts)
            updated.append(code)
    return added, updated


def _write_pool(pool):
    # 防御性写盘：内存池为空但磁盘已有数据 → 拒绝清空，避免 BOM/解析失败把观察池抹掉
    if not pool:
        existing_rows = 0
        if MASTER_CSV.exists():
            try:
                with MASTER_CSV.open(encoding="utf-8-sig") as _f:
                    # 跳过空行/标记行（日期分隔），只统计真实数据行
                    existing_rows = sum(1 for _r in csv.DictReader(_f)
                                        if (_r.get("代码") or "").strip()
                                        and not (_r.get("代码") or "").lstrip().startswith("#"))
            except Exception:
                existing_rows = 0
        if existing_rows > 0:
            print(f"[严重][拒绝写盘] 内存观察池为空，但磁盘 {MASTER_CSV.name} 现有 {existing_rows} 只。"
                  f"疑似 CSV 解析失败(BOM/编码)，已跳过写盘以防数据清空。请检查文件后重试。")
            return
    # 防冲突标记固化（双保险，与 load_master 的过滤配合）：pool 中任何"代码"字段
    # 以 git 冲突标记开头 → 判定为脏数据，拒绝写盘，防止把 <<<<<<< HEAD 等行 commit 入库。
    for _r in pool.values():
        _c = (_r.get("代码") or "").strip()
        if _c.startswith(("<<<<<<<", "=======", ">>>>>>>")):
            print(f"[严重][拒绝写盘] 观察池含 git 冲突标记行（{_c[:30]}...），"
                  f"疑似脏数据，已跳过写盘。请先 git reset/解决冲突后重试。")
            return
    MASTER_CSV.parent.mkdir(parents=True, exist_ok=True)
    with MASTER_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_COLS, extrasaction="ignore")
        w.writeheader()
        # 按「首次入选日期 + 首次入选时间」排序：不同日期批次按日期分组，
        # 同日内按首次入选时间（早的在前）；时间相同再按代码保证稳定顺序。
        # 注意：CSV 内不再写分隔行——任何非 28 列的行都会让 GitHub 拒绝按表格渲染
        # （Table tab 消失、退化成纯文本）。日期分组改由同目录的 .md 副本按天呈现。
        for r in sorted(pool.values(),
                        key=lambda x: (x.get("首次入选日期") or "",
                                       x.get("首次入选时间") or "",
                                       x.get("代码") or "")):
            w.writerow(r)


# ---------------- Markdown 副本（GitHub 网页友好） ----------------
# CSV 在 GitHub 上只要行数不一致就会退化成纯文本（Table tab 消失）。
# 为兼顾「Excel 打开 CSV 按列对齐」与「GitHub 网页按天分组 + 列对齐」，
# 额外生成一份 .md：按「首次入选日期」分组，每天一个二级标题 + 表格，
# 列对齐由 Markdown 渲染保证，与 CSV 列数无关。
MASTER_XLSX = OUTPUT_DIR / "观察池_累计.xlsx"

# xlsx 表格的列：与 CSV 完全一致（全部保留）。
# 日期分组用「单独整行」呈现（合并单元格 + 加粗，不涂色，简洁），
# 列本身与 CSV 相同，Excel / GitHub 网页打开的 xlsx 与 csv 内容一一对应。
XLSX_COLS = MASTER_COLS


def render_xlsx(pool):
    """把内存池渲染成按日期分组的 Excel 文件，写盘到 MASTER_XLSX。

    GitHub 网页原生支持 .xlsx 在线预览（无需下载，点开即渲染成表格），
    不像 markdown 表格那样受「列头竖排」窄屏优化影响，也不受 sanitize 限制，
    是所有格式里在 GitHub 上观感最接近本机 Excel 的方案。
    日期分组用单独整行呈现：跨列合并 + 加粗，不加背景色（简洁不花哨）。
    """
    if not pool:
        return
    rows = sorted(pool.values(),
                  key=lambda x: (x.get("首次入选日期") or "",
                                 x.get("首次入选时间") or "",
                                 x.get("代码") or ""))
    MASTER_XLSX.parent.mkdir(parents=True, exist_ok=True)

    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "观察池"

    n = len(XLSX_COLS)

    def write_row(values):
        ws.append(list(values))

    def row_idx():
        return ws.max_row

    prev_date = None
    for r in rows:
        d = (r.get("首次入选日期") or "").strip() or "未知日期"
        if d != prev_date:
            # 日期单独一整行：合并前 n 列 + 加粗（无背景色）
            write_row([f"📅 {d}"] + [""] * (n - 1))
            ws.merge_cells(start_row=row_idx(), start_column=1,
                           end_row=row_idx(), end_column=n)
            cell = ws.cell(row=row_idx(), column=1)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # 该日期表头（每个分组独立一份，深色字加粗）
            write_row(XLSX_COLS)
            for c in range(1, n + 1):
                hc = ws.cell(row=row_idx(), column=c)
                hc.font = Font(bold=True)
            prev_date = d
        write_row([str(r.get(c) or "").strip() for c in XLSX_COLS])

    # 列宽按内容自适应（上限 42，避免「入选扫描时间点」这类长字段撑爆）
    for c in range(1, n + 1):
        letter = get_column_letter(c)
        maxlen = max(len(str(ws.cell(row=rr, column=c).value or ""))
                     for rr in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 8), 42)
    # 冻结首行（滚动时表头常驻）
    ws.freeze_panes = "A2"

    wb.save(MASTER_XLSX)



def _scan_once(pool, force):
    """单次扫描并合并；返回是否有命中。"""
    now = now_shanghai()
    if not force and not in_trading_hours(now):
        print(f"[跳过] {now:%Y-%m-%d %H:%M} 非交易时段")
        return False
    print(f"[执行] {now:%Y-%m-%d %H:%M} 拉取妙想选股 ...")
    rows = run_screener()
    print(f"[结果] 本次命中 {len(rows)} 只")
    if not rows:
        return False
    # 基本面二次校验：TTM 净利<0 / 负债率>70% 硬拦截（不依赖妙想 prompt 自觉）
    rows = _fundamental_gate(rows)
    if not rows:
        print("[基本面] 候选全部被拦截，本轮无新增")
        return False
    added, updated = _merge_rows(pool, rows, now)
    print(f"[累计] 本次新增 {len(added)}: {added} | 刷新 {len(updated)}: {updated}")
    return True


# ---------------- T+1 次日表现跟踪 ----------------
# 思路：在股票「首次入选日的下一个交易日（T+1）」收盘后，抓取当天 OHLC + 11:30 午间价，
# 以昨收为基准换算涨跌幅，覆盖"上午涨下午跌"的日内路径（不只看收盘）。
# 数据源：腾讯自选股公开行情接口（qt.gtimg.cn 实时 + web.ifzq.gtimg.cn 分时），
# 全部以「元」为单位，无东方财富 push2 的 ×100 单位错位问题，且沙箱连通性更稳。

def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _http_get_json(url, params=None, timeout=6.0, retries=2):
    import time as _t
    last = None
    for attempt in range(retries):
        try:
            r = httpx.get(url, params=params, timeout=timeout,
                          headers={"User-Agent": "Mozilla/5.0", "Referer": "https://gu.qq.com/"})
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last = e
            if attempt < retries - 1:
                _t.sleep(1.5 * (attempt + 1))
    print(f"[跟踪] 请求失败 {url}: {last}")
    return None


def _http_get_text(url, timeout=6.0, retries=2):
    """腾讯 qt.gtimg.cn 返回纯文本（v_xxx="1~..."），需按文本解析。"""
    import time as _t
    last = None
    for attempt in range(retries):
        try:
            r = httpx.get(url, timeout=timeout,
                          headers={"User-Agent": "Mozilla/5.0", "Referer": "https://gu.qq.com/"})
            r.raise_for_status()
            return r.text
        except Exception as e:
            last = e
            if attempt < retries - 1:
                _t.sleep(1.5 * (attempt + 1))
    print(f"[跟踪] 请求失败 {url}: {last}")
    return None


def _tx_prefix(code, plate):
    """腾讯行情代码前缀：上交所/科创板用 sh，深交所/北交所用 sz。"""
    p = plate or ""
    if ("上交所" in p) or ("沪" in p) or ("科创" in p):
        return "sh"
    return "sz"  # 深交所、北交所及默认


def _next_trading_day(d):
    """返回 d 之后的第一个交易日（排除周末与法定节假日）。"""
    try:
        import chinese_calendar as cn
        nd = d
        while True:
            nd = nd + timedelta(days=1)
            if cn.is_workday(nd):
                return nd
    except Exception:
        pass
    nd = d
    while True:
        nd = nd + timedelta(days=1)
        if nd.weekday() < 5:
            return nd


def _fetch_live(code, plate):
    """腾讯实时行情（qt.gtimg.cn）：现价/最高/最低/今开/昨收，单位均为元。
    返回字段：price 现价, high 最高, low 最低, open 今开, prev_close 昨收,
             suspended 是否停牌（现价和今同时未变且成交量极低时判定）。
    停牌判定：现价==昨收==0 较罕见（退市股已被剔除），更常见是现价==昨收（不动）
    且今开==0；综合判断用「今开==0 且现价==昨收 且成交量<手」。"""
    pre = _tx_prefix(code, plate)
    text = _http_get_text(f"https://qt.gtimg.cn/q={pre}{code}")
    if not text:
        return None
    m = re.search(r'"([^"]*)"', text)
    if not m:
        return None
    f = m.group(1).split("~")
    if len(f) < 35:
        return None
    price = _num(f[3])        # 现价
    prev_close = _num(f[4])   # 昨收（作为 T+1 涨跌幅基准）
    open_p = _num(f[5])        # 今开
    high = _num(f[33])         # 最高
    low = _num(f[34])          # 最低
    volume = _num(f[6])        # 成交量(手)
    # 停牌典型特征：今开==0 且现价==昨收（无交易）
    suspended = (open_p == 0 and price is not None and price == prev_close
                  and (volume is None or volume == 0))
    return {
        "price": price,
        "prev_close": prev_close,
        "open": open_p,
        "high": high if not suspended else prev_close,  # 停牌时无日内波幅，取昨收
        "low": low if not suspended else prev_close,
        "suspended": suspended,
    }


def _snapshot_for_new(m):
    """新股首次入选：优先用腾讯实时价作买入基准（单位可靠、与 T+1/回填同源），
    腾讯不可达时回退妙想最新价（仅兜底，可能略偏）。返回 (入选价, 入选时涨跌幅%)。"""
    code = m.get("代码", "")
    plate = m.get("上市板块", "")
    live = None
    try:
        live = _fetch_live(code, plate)
    except Exception as e:  # 网络等异常，兜底妙想
        print(f"[快照] {code} 腾讯取价异常({type(e).__name__})，回退妙想")
    if live and live.get("price") and live.get("prev_close"):
        price = live["price"]
        prev = live["prev_close"]
        pct = round((price - prev) / prev * 100, 2) if prev else ""
        return price, pct
    # 兜底：妙想最新价（可能略偏，但保证有基准）
    print(f"[快照] {code} 腾讯取价失败，回退妙想最新价")
    return m.get("最新价(元)"), m.get("涨跌幅(%)")


def _fetch_mid_price(code, plate, target):
    """取目标日 11:30 的「原始价格（元）」，涨跌幅在调用处用同一 prev_close 计算，
    从源头避免单位错位。返回价格或 None。
    腾讯分时接口（web.ifzq.gtimg.cn）只返回最近一个交易日的数据；
    本函数仅用于 target == today 的收盘后跟踪，故天然命中目标日。"""
    pre = _tx_prefix(code, plate)
    url = f"https://web.ifzq.gtimg.cn/appstock/app/minute/query?code={pre}{code}"
    d = _http_get_json(url)
    if not d:
        return None
    node = d.get("data", {}).get(f"{pre}{code}", {}).get("data")
    if not isinstance(node, dict):
        return None
    # 交易日校验：仅当接口返回日 == 目标日才采信（防止跨日脏数据）
    if str(node.get("date")) != target.strftime("%Y%m%d"):
        return None
    rows = node.get("data") or []
    best = None
    for row in rows:
        parts = str(row).split()
        if len(parts) < 2:
            continue
        t = parts[0]              # "HHMM"（无冒号）
        if t <= "1130":
            best = _num(parts[1])
    if best is None:
        # 兜底：取 11:30 之后第一条（部分标的午间无独立点）
        for row in rows:
            parts = str(row).split()
            if len(parts) >= 2 and parts[0] >= "1130":
                return _num(parts[1])
        return None
    return best


def _classify_shape(open_p, mid_p, close_p, high_p, low_p):
    try:
        o = float(open_p); c = float(close_p)
        h = float(high_p) if high_p not in (None, "") else None
        l = float(low_p) if low_p not in (None, "") else None
    except (TypeError, ValueError):
        return "数据不足"
    m = None
    try:
        if mid_p not in (None, ""):
            m = float(mid_p)
    except (TypeError, ValueError):
        m = None
    # 上午涨、下午跌（用户核心场景），需要午间数据
    if m is not None and m > 0 and c < m - 0.5:
        return "冲高回落" if (h is not None and h > max(m, c) + 1) else "上午涨下午跌"
    if o > 0.5 and c > o:
        return "高开高走"
    if o > 0.5 and c < o:
        return "高开低走"
    if o < -0.5 and c > o:
        return "低开高走"
    if o < -0.5 and c < o:
        return "低开低走"
    return "震荡收涨" if c >= 0 else "震荡收跌"


def _fetch_day_minutes(code, plate, target):
    """用 day/query 回溯目标日（最近 ~5 交易日）的分时，返回原始价(元)字典：
    open/high/low/close/mid + prev_close(目标日昨收，取前一交易日收盘)。
    分时接口只能取最近几天，故仅对近期错过的 T+1 可补抓；返回 None 表示超窗口/无数据。"""
    pre = _tx_prefix(code, plate)
    url = f"https://web.ifzq.gtimg.cn/appstock/app/day/query?code={pre}{code}"
    d = _http_get_json(url)
    if not d:
        return None
    days = d.get("data", {}).get(f"{pre}{code}", {}).get("data") or []
    if not isinstance(days, list):
        return None
    tgt_s = target.strftime("%Y%m%d")
    for i, day in enumerate(days):
        if not isinstance(day, dict) or day.get("date") != tgt_s:
            continue
        rows = day.get("data") or []
        prices = []
        for row in rows:
            parts = str(row).split()
            if len(parts) < 2:
                continue
            p = _num(parts[1])
            if p is not None:
                prices.append((parts[0], p))
        if not prices:
            return None
        open_p = prices[0][1]
        close_p = prices[-1][1]
        high_p = max(p for _, p in prices)
        low_p = min(p for _, p in prices)
        mid = None
        for t, p in prices:
            if t <= "1130":
                mid = p
        prev_close = None
        if i + 1 < len(days):           # 列表按日期降序，下一项即前一交易日
            nxt = days[i + 1]
            if isinstance(nxt, dict):
                nrows = nxt.get("data") or []
                if nrows:
                    last = str(nrows[-1]).split()
                    if len(last) >= 2:
                        prev_close = _num(last[1])
        return {"open": open_p, "high": high_p, "low": low_p, "close": close_p,
                "mid": mid, "prev_close": prev_close}
    return None


def _apply_t1(ex, target, prev, o, h, l, c, m, status):
    """用统一 prev_close 把原始价换算为涨跌幅并写入 T1 字段，返回形态。"""
    def _pct(v):
        return round((float(v) - prev) / prev * 100, 2) if v is not None else ""
    open_p = _pct(o)
    high_p = _pct(h)
    low_p = _pct(l)
    close_p = _pct(c)
    mid_p = _pct(m)
    shape = _classify_shape(open_p, mid_p, close_p, high_p, low_p)
    ex["次日_跟踪状态"] = status
    ex["次日_跟踪日期"] = target.strftime("%Y-%m-%d")
    ex["次日_开盘涨跌幅"] = open_p
    ex["次日_午间涨跌幅"] = mid_p if mid_p is not None else ""
    ex["次日_收盘涨跌幅"] = close_p
    ex["次日_最高涨跌幅"] = high_p
    ex["次日_最低涨跌幅"] = low_p
    ex["次日_形态"] = shape
    # 策略口径收益：以「入选价」(信号出现时价格) 为买入基准，衡量跟信号买的真实盈亏
    # 而非个股相对昨收的日内表现。c/h/l 为 T+1 日原始价(元)，与入选价同单位。
    entry = _num(ex.get("入选价(元)"))
    if entry:
        def _strat(v):
            return round((float(v) - entry) / entry * 100, 2) if v is not None else ""
        ex["策略收益_次日收盘(%)"] = _strat(c)
        ex["策略收益_次日最高(%)"] = _strat(h)   # 次日盘中最高卖出的最佳情形
        ex["策略收益_次日最低(%)"] = _strat(l)   # 次日盘中最低卖出的回撤情形
    else:
        # 改版前入选的老样本无入选价，留空（不参与策略收益统计）
        ex["策略收益_次日收盘(%)"] = ""
        ex["策略收益_次日最高(%)"] = ""
        ex["策略收益_次日最低(%)"] = ""
    return shape


# 本次运行内 T+1 当天应抓但取数失败的股票（用于 CI ::error:: 告警，避免静默丢失）
T1_FETCH_FAILURES = []

def track_followups(pool, force):
    """对池中每只股票，在其 T+1 日记录次日表现；返回是否有变更。
    - 次日_当前涨跌幅：每次运行实时刷新（仅当 T+1==今天，盘中观察用）。
    - 次日_收盘涨跌幅等收盘字段：仅 T+1 日收盘后(after_close)抓取，标"已跟踪"后冻结；
      --force 不再提前触发（避免盘中实时价被误存为"收盘"）。
    - 错过的 T+1：用 day/query 回溯补抓(最近~5交易日)，标"已补抓"；超窗口才标"已过期"。
    取数失败记入 T1_FETCH_FAILURES 并打印 ::error::，使 CI 运行变红可见。"""
    global T1_FETCH_FAILURES
    T1_FETCH_FAILURES = []
    now = now_shanghai()
    today = now.date()
    after_close = now.time() >= datetime.strptime("15:00", "%H:%M").time()
    changed = False
    # 总时长预算：海外 runner 访问腾讯接口可能很慢，20+ 只股票逐只抓取
    # 若不加总预算，单次 track_followups 可拖 10+ 分钟（15s×3重试×20只=900s），
    # 超过 loop 的 deadline 保护，导致云端运行超长/卡死。预算内尽力抓，超时下轮再补。
    import time as _t
    budget = _t.monotonic() + float(os.environ.get("T1_FETCH_BUDGET", "75"))
    total = len(pool)
    processed = 0
    for key, ex in pool.items():
        processed += 1
        if _t.monotonic() > budget:
            print(f"[跟踪] 已达本轮 T+1 抓取预算，剩余 {total - processed + 1} 只留待下轮补抓")
            break
        # key 为「代码|入选日期」复合键；取数/打印一律用行内的真实代码
        code = ex.get("代码") or (key.split("|")[0] if key else "")
        first_s = (ex.get("首次入选日期") or "").strip()
        if not first_s:
            continue
        try:
            first_d = datetime.strptime(first_s, "%Y-%m-%d").date()
        except Exception:
            continue
        plate = ex.get("上市板块") or ""
        target = _next_trading_day(first_d)
        if target > today:
            continue                      # 还没到 T+1
        # T+1==今天：盘中即可得的字段立即抓取，不必等收盘。
        # 仅「收盘涨跌幅 + 形态」留到收盘后(after_close)固化，避免把盘中实时价误存为"收盘"。
        # 可盘中获取：当前/开盘/最高/最低(实时极值)、午间(11:30后)、跟踪日期、跟踪状态(跟踪中)。
        if target == today:
            if not (ex.get("次日_跟踪日期") or ""):
                ex["次日_跟踪日期"] = target.strftime("%Y-%m-%d")
                changed = True
            live0 = _fetch_live(code, plate)
            if live0 and live0.get("prev_close"):
                prev = live0["prev_close"]
                # 停牌判定仅在开盘后(>=9:30)有效：盘前(9:30前)无成交会被误判为停牌，
                # 此时不应标记为停牌，应跳过、等开盘后再识别真实停牌。
                if live0.get("suspended") and now.time() >= datetime.strptime("09:30", "%H:%M").time():
                    ex["次日_跟踪状态"] = "T+1停牌"
                    ex["次日_形态"] = "停牌无交易"
                    print(f"[停牌] {code} T+1 停牌，已标记（入选日 {first_s}，T+1 {target}）")
                    changed = True
                    continue
                price = live0.get("price")
                o = live0.get("open"); h = live0.get("high"); l = live0.get("low")

                def _p(v):
                    return round((float(v) - prev) / prev * 100, 2) if v is not None else ""
                # 当前涨跌幅：每次刷新（实时）；仅开盘后(>=9:30)抓取，盘前无真实现价不写
                if price and price > 0 and now.time() >= datetime.strptime("09:30", "%H:%M").time():
                    ex["次日_当前涨跌幅"] = _p(price); changed = True
                # 开盘涨跌幅：开盘后即固定，抓一次即可；仅在开盘后且为有效正值时抓
                # （盘前 open 可能为 0，避免锁定错误开盘涨跌幅）
                if o and o > 0 and not ex.get("次日_开盘涨跌幅"):
                    ex["次日_开盘涨跌幅"] = _p(o); changed = True
                # 最高/最低涨跌幅：日内实时极值，每次刷新；仅在有效正值时抓
                if h and h > 0:
                    ex["次日_最高涨跌幅"] = _p(h); changed = True
                if l and l > 0:
                    ex["次日_最低涨跌幅"] = _p(l); changed = True
                # 午间涨跌幅：11:30 之后可得，抓一次即固化
                if now.time() >= datetime.strptime("11:30", "%H:%M").time() and not ex.get("次日_午间涨跌幅"):
                    m = _fetch_mid_price(code, plate, target)
                    if m is not None:
                        ex["次日_午间涨跌幅"] = _p(m); changed = True
                # 跟踪状态：盘中标记"跟踪中"，收盘后改"已跟踪"（不再空白）
                if (ex.get("次日_跟踪状态") or "") not in ("已跟踪", "已补抓", "T+1停牌"):
                    ex["次日_跟踪状态"] = "跟踪中"; changed = True
        # 已抓过收盘 / 已确认停牌的：不重复处理
        if (ex.get("次日_跟踪状态") or "") in ("已跟踪", "已补抓", "T+1停牌"):
            continue
        if target < today:               # 错过窗口：回溯补抓
            kb = _fetch_day_minutes(code, plate, target)
            if kb and kb.get("prev_close"):
                # 补抓回看也需排除停牌日（分时 OHLC 全部相同 = 无实际成交）
                is_sus_bk = (kb["open"] == kb["close"] == kb["high"] == kb["low"]
                             and kb.get("open") is not None)
                if is_sus_bk:
                    ex["次日_跟踪状态"] = "T+1停牌"
                    ex["次日_跟踪日期"] = target.strftime("%Y-%m-%d")
                    ex["次日_形态"] = "停牌无交易"
                    print(f"[停牌] {code} T+1 停牌(补抓确认)，已标记")
                    changed = True
                    continue
                shape = _apply_t1(ex, target, kb["prev_close"], kb["open"],
                                  kb["high"], kb["low"], kb["close"], kb["mid"], "已补抓")
                ex["次日_当前涨跌幅"] = ex.get("次日_收盘涨跌幅", "")
                changed = True
                print(f"[补抓] {code} T+1 回溯成功: 开{ex['次日_开盘涨跌幅']} 午{ex['次日_午间涨跌幅']} 收{ex['次日_收盘涨跌幅']} 高{ex['次日_最高涨跌幅']} 低{ex['次日_最低涨跌幅']} [{shape}]")
            else:
                if (ex.get("次日_跟踪状态") or "") != "已过期":
                    ex["次日_跟踪状态"] = "已过期"
                    ex["次日_跟踪日期"] = target.strftime("%Y-%m-%d")
                    changed = True
                gap_days = (today - target).days
                hint = "（可能跨长假，day/query 仅保留近 5 个交易日）" if gap_days > 7 else ""
                print(f"::warning::T+1 回溯失败(超回溯窗口): {code} T+1={target} 距今{gap_days}天{hint}")
            continue
        # target == today：仅收盘后(after_close)抓取收盘字段（--force 不再绕过）
        if not after_close:
            continue
        live = _fetch_live(code, plate)
        if live and live.get("suspended"):
            # 停牌日由实时接口确认为停牌，直接标记无需兜底
            ex["次日_跟踪状态"] = "T+1停牌"
            ex["次日_形态"] = "停牌无交易"
            print(f"[停牌] {code} T+1 停牌(收盘后确认)，已标记")
            changed = True
            continue
        if not (live and live.get("prev_close")):
            kb = _fetch_day_minutes(code, plate, target)   # 实时失败，同源兜底
            if kb and kb.get("prev_close"):
                # 兜底成功时检查分时是否停牌（分时无数据或价格无变化）
                is_sus_bk = (kb["open"] == kb["close"] == kb["high"] == kb["low"]
                             and kb.get("open") is not None)
                if is_sus_bk:
                    ex["次日_跟踪状态"] = "T+1停牌"
                    ex["次日_形态"] = "停牌无交易"
                    print(f"[停牌] {code} T+1 停牌(day/query兜底确认)，已标记")
                    changed = True
                    continue
                shape = _apply_t1(ex, target, kb["prev_close"], kb["open"],
                                  kb["high"], kb["low"], kb["close"], kb["mid"], "已跟踪")
                changed = True
                print(f"[跟踪] {code} T+1 已记录(day/query兜底): {ex['次日_开盘涨跌幅']}/{ex['次日_午间涨跌幅']}/{ex['次日_收盘涨跌幅']} [{shape}]")
                continue
            print(f"::error::T+1 取数失败(当日应抓): {code} 取不到昨收/实时行情")
            T1_FETCH_FAILURES.append(code)
            continue
        prev = live["prev_close"]
        mid_price = _fetch_mid_price(code, plate, target)  # 11:30 原始价（元）
        shape = _apply_t1(ex, target, prev, live.get("open"), live.get("high"),
                          live.get("low"), live.get("price"), mid_price, "已跟踪")
        if ex["次日_午间涨跌幅"] == "":
            print(f"::warning::T+1 午间价缺失(其余字段正常): {code}")
        changed = True
        print(f"[跟踪] {code} T+1 已记录: 开{ex['次日_开盘涨跌幅']} 午{ex['次日_午间涨跌幅']} 收{ex['次日_收盘涨跌幅']} 高{ex['次日_最高涨跌幅']} 低{ex['次日_最低涨跌幅']} [{shape}]")
    return changed


def _report_t1_failures():
    """T+1 取数失败的收尾告警：让 CI 运行变红，但不影响已完成的 git 同步。"""
    if T1_FETCH_FAILURES:
        codes = ",".join(T1_FETCH_FAILURES)
        print(f"::error::T+1 跟踪存在 {len(T1_FETCH_FAILURES)} 只当日应抓却取数失败: {codes}")
        sys.exit(1)


def _install_log_tee():
    """把 stdout/stderr 同时写到控制台（若有）与 monitor.log，保证静默运行时仍可查日志。"""
    log_path = WORKSPACE / "monitor.log"
    try:
        # 日志轮转：超过 5MB 时把旧日志改名为 monitor.log.1（仅保留一份备份）
        try:
            if log_path.exists() and log_path.stat().st_size > 5 * 1024 * 1024:
                backup = log_path.with_suffix(".log.1")
                try:
                    if backup.exists():
                        backup.unlink()
                except OSError:
                    pass
                log_path.rename(backup)
        except Exception:
            pass
        _log_f = log_path.open("a", encoding="utf-8")
    except Exception:
        return
    class _Tee:
        def __init__(self, *streams):
            # 静默环境（pythonw / 计划任务）下 sys.stdout/stderr 可能为 None，
            # 过滤掉，保证 tee 只写日志文件也能正常工作。
            self._streams = [s for s in streams if s is not None]
            self.encoding = "utf-8"
            self.errors = "replace"
        def write(self, s):
            for st in self._streams:
                try:
                    st.write(s)
                except Exception:
                    pass
        def flush(self):
            for st in self._streams:
                try:
                    st.flush()
                except Exception:
                    pass
        def isatty(self):
            return False
        def writable(self):
            return True
        def readable(self):
            return False
    sys.stdout = _Tee(sys.stdout, _log_f)
    sys.stderr = _Tee(sys.stderr, _log_f)


def main():
    _install_log_tee()
    print(f"\n--- run {now_shanghai():%Y-%m-%d %H:%M:%S} ---", flush=True)
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true", help="忽略交易时段守卫")
    ap.add_argument("--loop", action="store_true",
                    help="单次运行内循环扫描（云端密集模式，约每2分钟一次）")
    ap.add_argument("--loop-interval", type=int, default=120,
                    help="循环间隔秒（默认120）")
    ap.add_argument("--loop-max-seconds", type=int, default=240,
                    help="单次运行最长持续秒（默认240，须<300避免与下次触发重叠）")
    ap.add_argument("--renew", action="store_true",
                    help="本地续期 EM_API_KEY（扫码一次，自动推送到仓库 Secret）")
    ap.add_argument("--no-push", action="store_true",
                    help="本地镜像模式：云端为主力时，本地只 pull 最新CSV供离线查看，不扫描不推送（避免与云端抢同一文件）")
    args = ap.parse_args()

    global _NO_PUSH
    _NO_PUSH = bool(args.no_push)

    if args.renew:
        cmd_renew()
        return

    if args.no_push:
        # 云端为主力写入时的本地镜像模式：只拉取最新 CSV 供离线查看，
        # 不扫描、不写盘、不推送，彻底避免与云端抢同一文件，也不产生未提交改动卡住 git pull。
        if not git_sync_before():
            print("[跳过] git pull 失败（代理/网络？），本轮镜像同步跳过，下次重试")
        else:
            print("[镜像] 云端为主力，本地仅 git pull 最新观察池，不扫描不推送。")
        return

    if not git_sync_before():
        # 拉取失败直接中止本轮：不 load_master、不扫描、不写盘。
        # 防止「基于旧数据扫描 → _write_pool 把 pull 冲突残留的标记行固化进 CSV」这类污染。
        print("[跳过] git pull 失败，本轮不执行（避免基于旧数据写盘污染观察池），下次运行自动重试")
        return
    pool = load_master()

    if args.loop:
        import time as _time
        interval = max(10, int(args.loop_interval))
        max_sec = min(int(args.loop_max_seconds), 295)
        deadline = _time.monotonic() + max_sec
        iterations = 0
        while _time.monotonic() < deadline:
            try:
                _scan_once(pool, args.force)
                track_followups(pool, args.force)  # 顺带处理 T+1 跟踪（收盘后抓取）
            except (MissingSecret, KeyExpired) as e:
                if isinstance(e, MissingSecret):
                    print(f"[致命] {e}")
                else:
                    _notify_key_expired()
                sys.exit(1)  # 让本次云端运行变红，明确暴露失败
            except Exception as e:
                print(f"[错误] {now_shanghai():%Y-%m-%d %H:%M} 单次迭代失败，跳过: {type(e).__name__}: {e}")
            iterations += 1
            if _time.monotonic() + interval < deadline:
                _time.sleep(interval)
            else:
                break
        _write_pool(pool)
        render_xlsx(pool)
        print(f"[循环结束] 共 {iterations} 次扫描尝试，观察池共 {len(pool)} 只")
        if not git_sync_after():  # 统一 commit+push 一次，与云端互不冲突
            print("::warning::git 同步失败，本次增量保留在本地，下次运行会自动重试")
        _report_t1_failures()
        return

    # 单次模式：先处理 T+1 跟踪（不受交易时段限制，收盘后也可跑），再视情况选股
    track_followups(pool, args.force)
    now = now_shanghai()
    if not args.force and not in_trading_hours(now):
        print(f"[跳过] 当前 {now:%Y-%m-%d %H:%M} 非交易时段，仅处理 T+1 跟踪，不执行选股扫描。")
    else:
        try:
            rows = run_screener()
            print(f"[结果] 本次命中 {len(rows)} 只")
            if rows:
                _merge_rows(pool, rows, now)
        except MissingSecret as e:
            print(f"[致命] {e}")
            sys.exit(1)
        except KeyExpired:
            _notify_key_expired()
            sys.exit(1)
        except Exception as e:
            print(f"[错误] 扫描失败: {type(e).__name__}: {e}")
    _write_pool(pool)
    render_xlsx(pool)
    print(f"[累计] 观察池共 {len(pool)} 只")
    if not git_sync_after():  # 把本地本次扫描合并回仓库，与云端互为补充
        print("::warning::git 同步失败，本次增量保留在本地，下次运行会自动重试")
    _report_t1_failures()


if __name__ == "__main__":
    main()
